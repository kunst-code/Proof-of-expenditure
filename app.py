# -*- coding: utf-8 -*-
import os
import io
import json
import re
from datetime import datetime

import streamlit as st
from supabase import create_client, Client
import google.genai as genai
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

st.set_page_config(page_title="사내 AI 지출증빙 시스템", page_icon="🧾", layout="centered")

SUPABASE_URL  = "https://dactmkqrckxzacrzihgc.supabase.co"
SUPABASE_ANON_KEY = "sb_publishable_PCH09T4MVJ7uCQhyll-VtQ_pB1MnRJm"

TEMPLATE_PATH = "template.xlsx"   # 양식 파일 (같은 폴더에 둘 것)

# ── 은행코드표 (주요) ──────────────────────────────────────────────
BANK_CODES = {
    "신한":   "088", "국민":   "004", "우리":   "020", "농협":   "011",
    "하나":   "081", "keb하나":"081", "기업":   "003", "우체국": "071",
    "카카오": "090", "카카오뱅크":"090", "토스":  "092", "토스뱅크":"092",
    "케이뱅크":"089", "부산":  "032", "광주":   "034", "제주":   "035",
    "전북":   "037", "경남":   "039", "새마을": "045", "신협":   "048",
    "sc":     "023", "씨티":   "053", "산업":   "002", "수협":   "007",
}

def load_gemini_key():
    try:
        supabase: Client = create_client(SUPABASE_URL.strip(), SUPABASE_ANON_KEY.strip())
        response = supabase.table("app_config").select("key_value").eq("key_name","GEMINI_API_KEY").execute()
        if response.data:
            return str(response.data[0]['key_value']).strip()
    except Exception as e:
        st.error(f"중앙 보안 서버 인증 실패: {e}")
    return None

def normalize_bank_code(raw: str) -> str:
    """은행명 또는 코드 문자열을 3자리 코드로 정규화"""
    raw = raw.strip()
    if re.fullmatch(r'\d{2,3}', raw):
        return raw.zfill(3)
    key = raw.lower().replace(" ","").replace("은행","")
    return BANK_CODES.get(key, raw)

def analyze_receipt(client, image: Image.Image) -> list[dict]:
    """이미지 → Gemini → JSON 파싱"""
    prompt = """
이 영수증/주문내역 이미지에서 계좌이체에 필요한 정보를 추출해 JSON 배열로만 응답하세요.
다른 텍스트 없이 순수 JSON만 출력하세요.

은행코드 참고(주요): 신한=088, 국민=004, 우리=020, 농협=011, KEB하나=081,
우체국=071, 카카오뱅크=090, 토스뱅크=092, 케이뱅크=089, 기업=003

JSON 형식:
[
  {
    "입금은행": "은행코드(숫자3자리, 예:088)",
    "입금계좌": "계좌번호(숫자만, 하이픈제거)",
    "고객관리성명": "수취인 이름(10자이내)",
    "입금액": "금액(숫자만, 쉼표없이, 부가세포함 합계)",
    "출금통장표시내용": "출금통장표시(7자이내한글)",
    "입금통장표시내용": "입금통장표시(7자이내한글)",
    "입금인코드": "",
    "비고": "품목/용도(10자이내)",
    "업체사용key": ""
  }
]

- 여러 건이면 여러 객체로 분리
- 정보 없으면 빈 문자열 ""
- 금액: 쉼표·원·VAT 글자 제거, 숫자만
- 계좌: 하이픈 없이 숫자만
"""
    response = client.models.generate_content(
        model='gemini-2.5-flash-lite',
        contents=[image, prompt]
    )
    raw = response.text.strip()
    raw = re.sub(r'^```json|^```|```$', '', raw, flags=re.MULTILINE).strip()
    data = json.loads(raw)
    # 은행코드 정규화
    for row in data:
        if row.get("입금은행"):
            row["입금은행"] = normalize_bank_code(row["입금은행"])
    return data

def build_excel(rows: list[dict]) -> bytes:
    """기존 양식(template.xlsx)의 '입력정보' 시트에 데이터 채워 반환"""
    FIELDS = ["입금은행","입금계좌","고객관리성명","입금액",
              "출금통장표시내용","입금통장표시내용","입금인코드","비고","업체사용key"]

    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb["입력정보"]
        # 기존 데이터 행 모두 삭제 (헤더 1행 유지)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        # 템플릿 없을 경우 새로 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "입력정보"
        headers = ["*입금은행","*입금계좌","고객관리성명","*입금액",
                   "출금통장표시내용","입금통장표시내용","입금인코드","비고","업체사용key"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, name="Arial")
            cell.fill = PatternFill("solid", start_color="D9E1F2")
            cell.alignment = Alignment(horizontal="center")

    # 헤더 행 스타일 복사용 (있을 경우)
    header_row_styles = []
    for col in range(1, 10):
        c = ws.cell(row=1, column=col)
        header_row_styles.append({
            "font": copy(c.font) if c.font else None,
            "fill": copy(c.fill) if c.fill else None,
            "alignment": copy(c.alignment) if c.alignment else None,
        })

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row in enumerate(rows, 2):
        for c_idx, field in enumerate(FIELDS, 1):
            val = row.get(field, "")
            # 입금액은 숫자로
            if field == "입금액" and val:
                try:
                    val = int(str(val).replace(",",""))
                except:
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center" if field in ["입금은행","입금인코드","업체사용key"] else "left")

    # 열 너비
    widths = [10, 22, 14, 13, 16, 16, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── UI ────────────────────────────────────────────────────────────
st.title("🧾 사내 AI 지출증빙 시스템")
st.markdown("영수증·주문 캡쳐 이미지를 업로드하면 AI가 분석 후 **입금요청 엑셀 양식**을 자동으로 채워 다운로드합니다.")
st.divider()

gemini_key = load_gemini_key()

if not gemini_key:
    st.warning("⚠️ 시스템 준비 중: 관리자 인증 동기화가 필요합니다.")
else:
    uploaded_files = st.file_uploader(
        "영수증 이미지를 선택하세요 (여러 장 가능)",
        type=["png","jpg","jpeg","webp"],
        accept_multiple_files=True
    )

    if uploaded_files:
        cols = st.columns(min(len(uploaded_files), 3))
        for i, f in enumerate(uploaded_files):
            with cols[i % 3]:
                st.image(f, caption=f.name, use_container_width=True)

        if st.button("🤖 AI 분석 후 엑셀 양식 자동 작성", type="primary"):
            all_rows = []
            errors = []

            with st.spinner("Gemini AI가 분석 중입니다..."):
                try:
                    ai_client = genai.Client(api_key=gemini_key)
                    for f in uploaded_files:
                        try:
                            img = Image.open(f)
                            rows = analyze_receipt(ai_client, img)
                            all_rows.extend(rows)
                            st.success(f"✅ {f.name} → {len(rows)}건 추출")
                        except Exception as e:
                            errors.append(f"{f.name}: {e}")
                            st.warning(f"⚠️ {f.name} 분석 실패: {e}")
                except Exception as e:
                    st.error(f"AI 초기화 실패: {e}")
                    all_rows = []

            if all_rows:
                st.divider()
                st.markdown(f"### 📊 추출 결과 — 총 **{len(all_rows)}건**")

                # 결과 테이블 미리보기
                import pandas as pd
                df = pd.DataFrame(all_rows)
                # 컬럼 순서 맞추기
                cols_order = ["입금은행","입금계좌","고객관리성명","입금액",
                              "출금통장표시내용","입금통장표시내용","비고"]
                df = df[[c for c in cols_order if c in df.columns]]
                st.dataframe(df, use_container_width=True)

                # 엑셀 생성 & 다운로드
                try:
                    excel_bytes = build_excel(all_rows)
                    today = datetime.now().strftime("%Y%m%d")
                    filename = f"입금요청_{today}.xlsx"
                    st.download_button(
                        label="⬇️ 입금요청 엑셀 다운로드",
                        data=excel_bytes,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    st.info(f"📁 파일명: `{filename}` — '입력정보' 시트에 데이터가 채워졌습니다.")
                except Exception as e:
                    st.error(f"엑셀 생성 실패: {e}")
